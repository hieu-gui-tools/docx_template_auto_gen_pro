"""Dialog Batch generate từ file Excel."""
from __future__ import annotations

from pathlib import Path

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QDialog, QFileDialog, QHBoxLayout, QLabel,
    QLineEdit, QMessageBox, QProgressBar, QPushButton,
    QTextEdit, QVBoxLayout,
)

STYLE = """
QDialog { background: #13172a; }
QLabel { color: #c8d0e8; font-size: 13px; }
QLineEdit {
    background: #1e2230; color: #e8eaf0;
    border: 1.5px solid #3a3f55; border-radius: 6px;
    padding: 5px 8px; font-size: 13px;
}
QTextEdit {
    background: #1a1f35; color: #c8d0e8;
    border: 1.5px solid #2a2f48; border-radius: 8px;
    font-family: Consolas, monospace; font-size: 12px;
}
QPushButton {
    border-radius: 6px; padding: 6px 16px;
    font-size: 13px; font-weight: bold;
}
QPushButton#primary { background: #2d3a8a; color: #aabbff; border: 1.5px solid #3d50c0; }
QPushButton#primary:hover { background: #3d50c0; }
QPushButton#success { background: #1a3a2a; color: #6bffaa; border: 1.5px solid #2a6040; }
QPushButton#success:hover { background: #2a6040; }
QPushButton#secondary { background: #1e2535; color: #8898cc; border: 1.5px solid #2a3050; }
QPushButton#secondary:hover { background: #2a3a5a; }
QProgressBar {
    background: #1a1f35; border: 1.5px solid #2a2f48;
    border-radius: 6px; height: 14px; text-align: center; color: #aabbff;
}
QProgressBar::chunk { background: #3d50c0; border-radius: 5px; }
"""


class BatchWorker(QThread):
    progress = Signal(int, int, str)   # current, total, msg
    finished = Signal(list)            # list of output paths

    def __init__(self, excel_path: str, docx_path: str, output_dir: str,
                 pattern: str, fields: list[dict], script_mod=None, template_name: str = ""):
        super().__init__()
        self.excel_path = excel_path
        self.docx_path  = docx_path
        self.output_dir = output_dir
        self.pattern    = pattern
        self.fields     = fields
        self.script_mod = script_mod
        self.template_name = template_name

    def run(self):
        import openpyxl
        from datetime import datetime
        from typing import Any
        from .template_engine import generate_document, FIELD_PATTERN
        import re

        results = []
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len(rows)
            for i, row_data in enumerate(rows):
                row_dict_typed: dict[str, Any] = {}
                row_dict_str: dict[str, str] = {}
                for j, v in enumerate(row_data):
                    if j >= len(headers):
                        break
                    k = headers[j]
                    row_dict_typed[k] = v
                    if isinstance(v, datetime):
                        if v.time() == datetime.min.time():
                            row_dict_str[k] = v.strftime("%d/%m/%Y")
                        else:
                            row_dict_str[k] = v.strftime("%d/%m/%Y %H:%M")
                    else:
                        row_dict_str[k] = str(v) if v is not None else ""
                
                # Build replacements
                replacements: dict[str, str] = {}
                field_values_typed: dict[str, Any] = {}
                
                # First pass: collect all non-script fields
                for f in self.fields:
                    col_name = f["name"]
                    raw_tag  = f["raw"]
                    ftype    = f["type"].lower()
                    if ftype == "script":
                        continue
                        
                    raw_val = row_dict_typed.get(col_name)
                    str_val = row_dict_str.get(col_name, "")
                    
                    if ftype == "text_lower": str_val = str_val.lower()
                    elif ftype == "text_upper": str_val = str_val.upper()
                    elif ftype == "text_capitalize": str_val = str_val.capitalize()
                    elif ftype == "text_title": str_val = str_val.title()
                    elif ftype == "integer":
                        try: str_val = str(int(float(str_val)))
                        except: pass
                    elif ftype in ("number", "float"):
                        try: str_val = f"{float(str_val):.4f}".rstrip("0").rstrip(".")
                        except: pass
                        
                    replacements[raw_tag] = str_val
                    
                    # Store typed value
                    typed_val = raw_val
                    if ftype == "integer":
                        try: typed_val = int(float(str_val))
                        except: typed_val = raw_val
                    elif ftype in ("number", "float"):
                        try: typed_val = float(str_val)
                        except: typed_val = raw_val
                    elif ftype == "date":
                        if isinstance(raw_val, datetime): typed_val = raw_val.date()
                        else:
                            try: typed_val = datetime.strptime(str_val, "%d/%m/%Y").date()
                            except: pass
                    elif ftype in ("datetime", "date_time"):
                        if isinstance(raw_val, datetime): typed_val = raw_val
                        else:
                            try: typed_val = datetime.strptime(str_val, "%d/%m/%Y %H:%M")
                            except: pass
                    else:
                        typed_val = str_val
                        
                    field_values_typed[col_name] = typed_val

                # Second pass: compute script fields
                import inspect
                for f in self.fields:
                    col_name = f["name"]
                    raw_tag  = f["raw"]
                    ftype    = f["type"].lower()
                    if ftype != "script":
                        continue
                        
                    # 1. Override from Excel if present
                    if col_name in headers and row_dict_typed.get(col_name) is not None and str(row_dict_typed.get(col_name)).strip() != "":
                        value = row_dict_typed[col_name]
                        replacements[raw_tag] = str(row_dict_str.get(col_name, ""))
                        field_values_typed[col_name] = value
                        continue
                        
                    # 2. Compute via script
                    if self.script_mod:
                        try:
                            from .template_engine import get_script_args_info, call_script_func
                            arg_info = get_script_args_info(self.script_mod, col_name)
                            kwargs = {}
                            for arg, ann in arg_info.items():
                                if arg in field_values_typed:
                                    kwargs[arg] = field_values_typed[arg]
                                else:
                                    raw_val = row_dict_typed.get(arg)
                                    str_val = row_dict_str.get(arg, "")
                                    
                                    ann_str = ""
                                    if ann != inspect.Parameter.empty:
                                        if isinstance(ann, str):
                                            ann_str = ann.lower()
                                        elif hasattr(ann, "__name__"):
                                            ann_str = ann.__name__.lower()
                                            
                                    val = raw_val
                                    if ann_str == "int":
                                        try: val = int(float(str_val))
                                        except: val = raw_val
                                    elif ann_str in ("float", "number"):
                                        try: val = float(str_val)
                                        except: val = raw_val
                                    elif ann_str == "str":
                                        val = str_val
                                    elif ann_str in ("datetime", "date"):
                                        if isinstance(raw_val, datetime): val = raw_val
                                        else:
                                            try: val = datetime.strptime(str_val, "%d/%m/%Y %H:%M")
                                            except: pass
                                    kwargs[arg] = val
                            value = call_script_func(self.script_mod, col_name, kwargs)
                        except Exception as e:
                            value = f"[Lỗi: {e}]"
                    else:
                        value = row_dict_str.get(col_name, "")
                    replacements[raw_tag] = str(value)
                    field_values_typed[col_name] = value

                # Filename
                now = datetime.now()
                fname = self.pattern
                for k, v in row_dict_str.items():
                    fname = fname.replace(f"{{{k}}}", v)
                fname = fname.replace("{template_name}", self.template_name)
                fname = fname.replace("{date}", now.strftime("%Y%m%d"))
                fname = fname.replace("{datetime}", now.strftime("%Y%m%d_%H%M"))
                fname = re.sub(r"[\\/:*?\"<>|]", "_", fname)
                base_fname = fname[:-5] if fname.endswith(".docx") else fname
                out_path = Path(self.output_dir) / f"{base_fname}.docx"
                counter = 1
                while out_path.exists():
                    out_path = Path(self.output_dir) / f"{base_fname}_{counter}.docx"
                    counter += 1
                out_path_str = str(out_path)
                
                try:
                    generate_document(self.docx_path, replacements, out_path_str)
                    results.append(out_path_str)
                    self.progress.emit(i + 1, total, f"✅ {out_path.name}")
                except Exception as e:
                    self.progress.emit(i + 1, total, f"❌ {fname}: {e}")
        except Exception as e:
            self.progress.emit(0, 0, f"Lỗi đọc Excel: {e}")
        self.finished.emit(results)


class BatchDialog(QDialog):
    def __init__(self, docx_path: str, output_dir: str, pattern: str,
                 fields: list[dict], script_mod=None, template_name: str = "", parent=None):
        super().__init__(parent)
        self.docx_path  = docx_path
        self.output_dir = output_dir
        self.pattern    = pattern
        self.fields     = fields
        self.script_mod = script_mod
        self.template_name = template_name
        self.worker     = None

        self.setWindowTitle("Batch Generate – Tạo nhiều file")
        self.setMinimumSize(620, 480)
        self.setStyleSheet(STYLE)
        self._build_ui()

    def _build_ui(self):
        v = QVBoxLayout(self)
        v.setSpacing(12)
        v.setContentsMargins(18, 18, 18, 18)

        title = QLabel("📊  Batch tạo file từ Excel")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #aabbff;")
        v.addWidget(title)

        # Excel picker
        h = QHBoxLayout()
        self.excel_edit = QLineEdit()
        self.excel_edit.setPlaceholderText("Chọn file Excel (.xlsx)…")
        self.excel_edit.setReadOnly(True)
        btn_pick = QPushButton("📂 Chọn file")
        btn_pick.setObjectName("primary")
        btn_pick.clicked.connect(self._pick_excel)
        h.addWidget(self.excel_edit)
        h.addWidget(btn_pick)
        v.addLayout(h)

        # Format hint
        normal_fields = [f["name"] for f in self.fields if f["type"].lower() != "script"]
        script_fields = [f["name"] for f in self.fields if f["type"].lower() == "script"]
        
        required_cols = set(normal_fields)
        if self.script_mod:
            try:
                from .template_engine import get_script_args_info
                for sf in script_fields:
                    args = get_script_args_info(self.script_mod, sf)
                    for arg in args:
                        if arg not in normal_fields:
                            required_cols.add(arg)
            except Exception:
                pass
                
        req_list = sorted(list(required_cols))
        
        hint = QLabel(
            f"ℹ️  File Excel cần có hàng đầu tiên là tên cột tương ứng.\n"
            f"Các cột cần thiết: {', '.join(req_list) if req_list else 'Không có'}\n\n"
            f"* Mẹo: Các trường type script sẽ tự động chạy tính toán. Nếu file Excel của bạn khai báo cột có cùng tên với trường script, giá trị trong cột đó sẽ ghi đè và ưu tiên sử dụng thay vì chạy script."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #6880a8; font-size: 11px;")
        v.addWidget(hint)

        # Progress
        self.progress = QProgressBar()
        self.progress.setValue(0)
        v.addWidget(self.progress)

        # Log
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        v.addWidget(self.log)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_run = QPushButton("▶  Bắt đầu tạo")
        self.btn_run.setObjectName("success")
        self.btn_run.clicked.connect(self._run_batch)
        btn_close = QPushButton("✖  Đóng")
        btn_close.setObjectName("secondary")
        btn_close.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_run)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        v.addLayout(btn_row)

    def _pick_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file Excel", "", "Excel Files (*.xlsx *.xls)")
        if path:
            self.excel_edit.setText(path)

    def _run_batch(self):
        excel = self.excel_edit.text()
        if not excel:
            QMessageBox.warning(self, "Thiếu file", "Vui lòng chọn file Excel.")
            return
        self.log.clear()
        self.progress.setValue(0)
        self.btn_run.setEnabled(False)
        self.worker = BatchWorker(excel, self.docx_path, self.output_dir,
                                  self.pattern, self.fields, self.script_mod, self.template_name)
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_finished)
        self.worker.start()

    def _on_progress(self, cur: int, total: int, msg: str):
        if total > 0:
            self.progress.setMaximum(total)
            self.progress.setValue(cur)
        self.log.append(msg)

    def _on_finished(self, results: list):
        self.btn_run.setEnabled(True)
        self.log.append(f"\n✅ Hoàn thành! Đã tạo {len(results)} file.")
        if results:
            import subprocess, os
            subprocess.Popen(f'explorer /select,"{results[0]}"')
